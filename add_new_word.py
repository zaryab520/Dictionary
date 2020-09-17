import xlsxwriter
import openpyxl


def newdic():
    
    ''' newdic() stands for new dictionary, it adds new word to the dictionary. The function
        asks Dari and Pashto translation of the word from the user.

        test5.xlsx is the database that contains words.
     '''

    newword = input('Please enter the word you want to add: ')
    NewWord = newword.upper()

    
    wb = openpyxl.load_workbook('dictionary.xlsx')
    ws = wb.active
    row = ws.max_row
    column = ws.max_column

    # using for loop it iterates all the rows in Database and matches the word
    for row in ws.iter_rows(ws.min_row, ws.max_row):
        for cell in row:
            if cell.value == NewWord:
                print('This word already exists')
                newdic()
                break

    Dari_trans = input('Enter the Dari Translation of the word: ')
    pashot_trans = input('Enter the Pashto Translation of the word: ')

    wb = openpyxl.load_workbook('dictionary.xlsx')
    ws = wb.active
    row = ws.max_row
    column = ws.max_column

    dic_list = [NewWord,Dari_trans,pashot_trans]
    for value in dic_list:
        ws.cell(row+1, column-5).value = value
        column+=1
    wb.save('dictionary.xlsx')
