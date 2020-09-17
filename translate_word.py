import xlsxwriter
import openpyxl
import add_new_word



print('Dictionary'.center(100,'='))
print(''' 
Welcome to Dictionary application you can use to translate English word to Dari & Pashto 
''')

rowpos = 0
colpos = 0

def dic():

    """ dic() function first ask the user for the English word, then it prompt for the 
        language that user needs the word to be translated to (Dari/Pashto).

        It also provide to users the ability to add new translation to dictionary if it doesnt
        exists.
    
     """
    english_word = input('Please enter the word to Translate: ')
    English_Word = english_word.upper()

    while True:
        d1a = input ("Please Choose the Language to Translate: A) Dari. B) Pashto. [A/B]? : ")
    
        if d1a in ['A', 'B']:
            break

    if d1a == "A": 
        wb = openpyxl.load_workbook('dictionary.xlsx') 
        ws = wb['emailcontent']
        flag = False
        for row in ws.iter_rows(ws.min_row, ws.max_row):
            for cell in row:
                if cell.value == English_Word:
                    rowpos = cell.row
                    colpos = cell.column
                    print(ws.cell(rowpos, colpos+1).value)
                    flag = True
                    break
            if flag:
                break
        else:
            userval = input('No such word found. Do you want to add the word in the Dictionary? (y/n): ')
            if userval == 'y':
                add_new_word.newdic()
            else:
                print('thanks for using our product')




    elif d1a == "B": 
        wb = openpyxl.load_workbook('dictionary.xlsx')
        ws = wb['emailcontent']
        flag = False
        for row in ws.iter_rows(ws.min_row, ws.max_row):
            for cell in row:
                if cell.value == English_Word:
                    rowpos = cell.row 
                    colpos = cell.column
                    print(ws.cell(rowpos, colpos+2).value)
                    flag = True
                    break
            if flag:
                break
        else:
            userval = input('No such word found. Do you want to add the word in the Dictionary? (y/n): ')
            if userval == 'y':
                
                add_new_word.newdic()
                
            else:
                print('Thank you for using our product')                   
dic()




